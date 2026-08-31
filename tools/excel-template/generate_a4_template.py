#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sinh MẪU EXCEL A4 gửi khách (chị Nguyễn Thu Phương) điền:
  - Danh sách Đại lý / CTV  (model Prisma `Dealer`)
  - Map nhóm Zalo -> Đại lý (model Prisma `Group`)

Đây là VIỆC KẾ TIẾP #1 (11/07/2026) — hạng mục Phase 3 DUY NHẤT không bị khách chặn:
"mẫu file gửi khách soạn ngay không cần chờ" (docs/phat-trien/ke-hoach/nen-tang.md §1.3, cổng A4).

Cột + dropdown khớp ĐÚNG enum trong apps/api/prisma/schema.prisma để importer sau này
(read-excel-file, xem docs/phat-trien/ke-hoach/nen-tang.md §1.3) đọc vào không cần suy đoán.
Bảng ánh xạ nhãn tiếng Việt -> giá trị enum: xem tools/excel-template/README.md.

Chạy:
    python tools/excel-template/generate_a4_template.py
Đầu ra:
    docs/khach-hang/ultty/trao-doi/a4-dai-ly-map-nhom-ultty.xlsx
Cần: openpyxl (pip install openpyxl). KHÔNG phải dependency runtime — chỉ là công cụ soạn mẫu.
"""

from __future__ import annotations

import sys
from pathlib import Path

# Console Windows mặc định cp1252 -> in tiếng Việt sẽ lỗi. Ép stdout/stderr UTF-8.
for _stream in (sys.stdout, sys.stderr):
    reconfigure = getattr(_stream, "reconfigure", None)
    if reconfigure is not None:
        reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

VERSION = "13/07/2026"

# --- Ánh xạ nhãn hiển thị (khách chọn) -> giá trị enum (importer ghi DB). NGUỒN SỰ THẬT DUY NHẤT ---
# Dealer.tier : DealerTier
TIER_LABEL_TO_ENUM = {
    "Đại lý": "dai_ly",
    "CTV (Cộng tác viên)": "ctv",
}
# Dealer.defaultPolicy : PolicyType
POLICY_LABEL_TO_ENUM = {
    "Công nợ 30 ngày": "cong_no_30",
    "Công nợ 45 ngày": "cong_no_45",
    "Ký gửi": "ky_gui",
    "Thanh toán ngay": "thanh_toan_ngay",
    "COD (thu hộ khi giao)": "cod",
}
POLICY_EXPLAIN = {
    "Công nợ 30 ngày": "Đại lý trả sau, tối đa 30 ngày.",
    "Công nợ 45 ngày": "Đại lý trả sau, tối đa 45 ngày.",
    "Ký gửi": "Cuối tháng chốt số đã bán rồi lên đơn (+VAT).",
    "Thanh toán ngay": "Trả ngay khi đặt (thường CTV nhỏ).",
    "COD (thu hộ khi giao)": "Thu hộ tiền khi giao, có phí thu hộ theo biểu mẫu.",
}

# --- Palette ---
NAVY = "1F3A5F"      # nền tiêu đề
STEEL = "34618E"     # nền phụ / tiêu đề bảng chú giải
CREAM = "F5F1E8"     # nền panel hướng dẫn
LIGHT = "EAF0F6"     # sọc dòng
WHITE = "FFFFFF"
INK = "1B2733"       # chữ đậm
MUTED = "5A6B7B"     # chữ phụ

THIN = Side(style="thin", color="C7D0DA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _header_cell(ws: Worksheet, col: int, text: str, width: float) -> None:
    cell = ws.cell(row=1, column=col, value=text)
    cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    cell.fill = _fill(NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(col)].width = width


def _data_cell(ws: Worksheet, row: int, col: int, value, *, wrap: bool = False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", size=11, color=INK)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    cell.border = BORDER
    if row % 2 == 1:  # sọc nhẹ cho dòng lẻ (row 3,5,...) để dễ đọc
        cell.fill = _fill(LIGHT)


def _list_validation(labels: list[str], prompt: str) -> DataValidation:
    """Dropdown danh sách cố định (inline). Không nhãn nào được chứa dấu phẩy."""
    formula = '"' + ",".join(labels) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
    dv.error = "Vui lòng chọn đúng 1 giá trị trong danh sách."
    dv.errorTitle = "Giá trị không hợp lệ"
    dv.prompt = prompt
    dv.promptTitle = "Chọn từ danh sách"
    return dv


# ----------------------------------------------------------------------------
# Sheet 1 — Đại lý & CTV
# ----------------------------------------------------------------------------
DEALER_COLS = [
    ("Tên đại lý / CTV (*)", 30),
    ("Cấp (*)", 20),
    ("Chính sách mặc định (*)", 24),
    ("Số điện thoại", 16),
    ("Tên gọi tắt / viết tắt\n(cách nhau dấu phẩy)", 30),
    ("Mã đại lý\n(để trống nếu chưa có)", 20),
]
# DÒNG VÍ DỤ — TỔNG HỢP, KHÔNG phải dữ liệu của khách.
#
# Trước 30/08/2026 ba dòng này là đại lý THẬT lấy từ khảo sát, và bản .xlsx sinh ra được commit
# vào một repo PUBLIC. Mẫu gửi khách chỉ cần cho thấy ĐỊNH DẠNG của một dòng hợp lệ; nó không
# cần biết khách có những đại lý nào. Xem docs/phat-trien/van-hanh/nguon-khach-hang.md.
DEALER_ROWS = [
    ["Đại lý mẫu A", "Đại lý", "Công nợ 30 ngày", "", "dai ly mau a, dlma", ""],
    ["Đại lý mẫu B", "Đại lý", "Công nợ 45 ngày", "", "dai ly mau b, dlmb", ""],
    ["CTV mẫu C", "CTV (Cộng tác viên)", "Thanh toán ngay", "", "ctv mau c, ctvmc", ""],
]
DEALER_SHEET = "1. Đại lý & CTV"
DEALER_LAST_ROW = 500  # phạm vi áp dropdown + named range cho dropdown ở sheet 2


def build_dealer_sheet(ws: Worksheet) -> None:
    ws.sheet_properties.tabColor = NAVY
    for idx, (title, width) in enumerate(DEALER_COLS, start=1):
        _header_cell(ws, idx, title, width)
    ws.row_dimensions[1].height = 34

    for r, row in enumerate(DEALER_ROWS, start=2):
        for c, value in enumerate(row, start=1):
            _data_cell(ws, r, c, value, wrap=(c == 5))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DEALER_COLS))}1"

    tier_dv = _list_validation(list(TIER_LABEL_TO_ENUM), "Chọn: Đại lý hoặc CTV.")
    policy_dv = _list_validation(
        list(POLICY_LABEL_TO_ENUM), "Chọn chính sách thanh toán mặc định của đại lý này."
    )
    ws.add_data_validation(tier_dv)
    ws.add_data_validation(policy_dv)
    tier_dv.add(f"B2:B{DEALER_LAST_ROW}")
    policy_dv.add(f"C2:C{DEALER_LAST_ROW}")


# ----------------------------------------------------------------------------
# Sheet 2 — Map nhóm Zalo -> Đại lý
# ----------------------------------------------------------------------------
GROUP_COLS = [
    ("Tên nhóm Zalo (*)", 34),
    ("Thuộc đại lý / CTV (*)", 30),
    ("Chi nhánh", 14),
    ("Chat ID nhóm\n(kỹ thuật — để trống, hệ thống điền)", 34),
]
# Cột Chat ID để TRỐNG — đúng như hướng dẫn ở sheet 1 dặn khách. Bản trước tự điền sẵn hai chat
# ID nhóm Zalo THẬT vào đây, tức mẫu vừa mâu thuẫn với chính hướng dẫn của nó, vừa mang định danh
# nhóm của khách vào file đem đi phát tán.
GROUP_ROWS = [
    ["Nhóm đại lý mẫu A", "Đại lý mẫu A", "HN", ""],
    ["Nhóm đại lý mẫu B", "Đại lý mẫu B", "TN", ""],
]
GROUP_SHEET = "2. Map nhóm Zalo"
DEALER_NAME_RANGE = "DanhSachDaiLy"


def build_group_sheet(ws: Worksheet, wb: Workbook) -> None:
    ws.sheet_properties.tabColor = STEEL
    for idx, (title, width) in enumerate(GROUP_COLS, start=1):
        _header_cell(ws, idx, title, width)
    ws.row_dimensions[1].height = 34

    for r, row in enumerate(GROUP_ROWS, start=2):
        for c, value in enumerate(row, start=1):
            # Chat ID là chuỗi số dài -> ép text để Excel không cắt/đổi số mũ.
            if c == 4 and value:
                cell = ws.cell(row=r, column=c, value=str(value))
                cell.number_format = "@"
                cell.font = Font(name="Calibri", size=11, color=INK)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = BORDER
                if r % 2 == 1:
                    cell.fill = _fill(LIGHT)
            else:
                _data_cell(ws, r, c, value)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(GROUP_COLS))}1"

    # Named range trỏ cột tên đại lý ở sheet 1 -> dropdown "Thuộc đại lý" luôn đồng bộ.
    ref = f"'{DEALER_SHEET}'!$A$2:$A${DEALER_LAST_ROW}"
    wb.defined_names[DEALER_NAME_RANGE] = DefinedName(DEALER_NAME_RANGE, attr_text=ref)
    dealer_dv = DataValidation(
        type="list", formula1=f"={DEALER_NAME_RANGE}", allow_blank=True, showErrorMessage=True
    )
    dealer_dv.error = "Tên đại lý phải khớp cột 'Tên đại lý / CTV' ở sheet 1."
    dealer_dv.errorTitle = "Đại lý chưa có trong danh sách"
    dealer_dv.prompt = "Chọn đại lý (lấy từ sheet '1. Đại lý & CTV')."
    dealer_dv.promptTitle = "Nhóm này của đại lý nào?"
    ws.add_data_validation(dealer_dv)
    dealer_dv.add("B2:B500")

    # Chú thích ngay dưới bảng.
    note_row = 6
    note = ws.cell(
        row=note_row,
        column=1,
        value="Ghi chú: mỗi nhóm Zalo chỉ thuộc 1 đại lý. Cột Chat ID để trống — kỹ thuật sẽ điền "
        "sau khi nhóm gửi tin đầu tiên (hệ thống tự bắt ID nhóm).",
    )
    note.font = Font(name="Calibri", size=10, italic=True, color=MUTED)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=4)
    ws.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height = 30


# ----------------------------------------------------------------------------
# Sheet Hướng dẫn
# ----------------------------------------------------------------------------
def _write(ws: Worksheet, row: int, text: str, *, bold=False, size=11, color=INK, italic=False):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    return cell


def build_guide_sheet(ws: Worksheet) -> None:
    ws.sheet_properties.tabColor = "9A7B4F"
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 40
    ws.sheet_view.showGridLines = False

    # Banner tiêu đề
    ws.merge_cells("A1:B2")
    banner = ws.cell(row=1, column=1, value="U ULTTY VIỆT NAM\nMẫu thu thập Đại lý & Map nhóm Zalo")
    banner.font = Font(name="Calibri", bold=True, size=15, color=WHITE)
    banner.fill = _fill(NAVY)
    banner.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 26

    r = 4
    _write(ws, r, "Mục đích", bold=True, size=12, color=NAVY); r += 1
    _write(
        ws, r,
        "File giúp đội Sale khai báo danh sách đại lý/CTV và ghép mỗi nhóm Zalo với đúng đại lý. "
        "Đây là bước để hệ thống áp đúng chính sách & giá cho từng nhóm khi đọc tin đặt hàng.",
    ); ws.row_dimensions[r].height = 44; r += 2

    _write(ws, r, "Cách điền", bold=True, size=12, color=NAVY); r += 1
    for line in [
        "1) Mở sheet “1. Đại lý & CTV”: mỗi dòng 1 đại lý/CTV. Các cột có dấu (*) là bắt buộc.",
        "2) Cột “Cấp” và “Chính sách mặc định”: bấm vào ô sẽ hiện danh sách để chọn (không gõ tay).",
        "3) Mở sheet “2. Map nhóm Zalo”: mỗi dòng 1 nhóm Zalo, chọn nhóm đó thuộc đại lý nào.",
        "4) Cột “Chat ID” để TRỐNG — bộ phận kỹ thuật điền sau, chị không cần biết ID.",
        "5) 3 dòng đại lý + 2 nhóm điền sẵn chỉ là VÍ DỤ về cách ghi — xoá đi rồi điền dữ liệu thật của mình.",
        "6) Đừng xóa dòng tiêu đề (dòng 1) ở 2 sheet dữ liệu.",
    ]:
        _write(ws, r, line); ws.row_dimensions[r].height = 30; r += 1
    r += 1

    # Bảng chú giải "Cấp"
    _write(ws, r, "Giá trị chọn — cột “Cấp”", bold=True, size=12, color=NAVY); r += 1
    r = _legend_table(ws, r, ["Chọn trong file", "Ý nghĩa"],
                      [(k, "Đại lý cấp 1" if v == "dai_ly" else "Cộng tác viên bán lẻ")
                       for k, v in TIER_LABEL_TO_ENUM.items()])
    r += 1

    # Bảng chú giải "Chính sách"
    _write(ws, r, "Giá trị chọn — cột “Chính sách mặc định”", bold=True, size=12, color=NAVY); r += 1
    r = _legend_table(ws, r, ["Chọn trong file", "Ý nghĩa"],
                      [(k, POLICY_EXPLAIN[k]) for k in POLICY_LABEL_TO_ENUM])
    r += 1

    _write(ws, r, f"Phiên bản {VERSION} · NetViet soạn cho U Ultty Việt Nam · liên hệ điền: chị Nguyễn Thu Phương",
           italic=True, size=10, color=MUTED)


def _legend_table(ws: Worksheet, start_row: int, headers: list[str], rows: list[tuple[str, str]]) -> int:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        cell.fill = _fill(STEEL)
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")
    r = start_row + 1
    for left, right in rows:
        lc = ws.cell(row=r, column=1, value=left)
        rc = ws.cell(row=r, column=2, value=right)
        for cell in (lc, rc):
            cell.font = Font(name="Calibri", size=11, color=INK)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = _fill(CREAM)
        ws.row_dimensions[r].height = 26
        r += 1
    return r


def main() -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "Hướng dẫn"
    build_guide_sheet(guide)

    dealer_ws = wb.create_sheet(DEALER_SHEET)
    build_dealer_sheet(dealer_ws)

    group_ws = wb.create_sheet(GROUP_SHEET)
    build_group_sheet(group_ws, wb)

    # `docs/mau/` khong con ton tai sau khi sap xep lai tai lieu theo khach — script van ghi vao
    # do nen chay lai KHONG tai sinh duoc tep dang duoc tham chieu. Do la ly do ban .xlsx trong
    # repo trot lech khoi generator ma khong ai thay.
    #
    # Dau ra KHONG duoc commit (.gitignore): day la mot ban build de gui khach, con byte goc va
    # ban khach tra ve deu thuoc ve kho rieng.
    out_dir = Path(__file__).resolve().parents[2] / "docs" / "khach-hang" / "ultty" / "trao-doi"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "a4-dai-ly-map-nhom-ultty.xlsx"
    wb.save(out_path)
    print(f"Đã tạo: {out_path}")
    print(f"  Sheet: {wb.sheetnames}")
    print(f"  Đại lý mẫu: {len(DEALER_ROWS)} · Nhóm mẫu: {len(GROUP_ROWS)}")


if __name__ == "__main__":
    main()
